"""Validate the owner-provided IC FlashMap reference snapshot."""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FLASHMAP_ROOT = ROOT / "refcode" / "flashmap" / "IC FlashMap"
COMBINER_SHA256 = "ed6b58289cc780f73d36b831f5424cef44ad93187ba7518d36df6a77ad0c76bf"

REQUIRED_SHEETS = {
    "TP Overview",
    "51923 TP Flashmap",
    "51926 TP Flashmap",
    "51927 TP Flashmap",
    "51929&51932 TP Flashmap",
    "51930 TP Flashmap",
    "51950&51951 TP Flashmap",
    "51950 DP Perspective",
}

REQUIRED_FILES = {
    "IC_FlashMap.xlsx",
    "combiner_1.13.0/Combiner.exe",
    "postbuild/PostbuildSetup_51920_1.3.1.bat",
    "postbuild/PostbuildSetup_51923_1.4.1.bat",
    "postbuild/PostbuildSetup_51926_2.0.0.bat",
    "postbuild/PostbuildSetup_51927_1.4.1.bat",
    "postbuild/PostbuildSetup_51930_2.0.0.bat",
    "postbuild/PostbuildSetup_51931_1.3.0.bat",
    "postbuild/PostbuildSetup_51932_2.0.0.bat",
    "postbuild/PostbuildSetup_51950_2.0.0.bat",
    "mmap.h/51920_1.3.1_mmap.h",
    "mmap.h/51923_1.4.1_mmap.h",
    "mmap.h/51926_2.0.0_mmap.h",
    "mmap.h/51927_1.4.1_mmap.h",
    "mmap.h/51929&51932_2.0.0_mmap.h",
    "mmap.h/51930_2.0.0_mmap.h",
    "mmap.h/51931_1.3.0_mmap.h",
    "mmap.h/51950_2.0.0_mmap.h",
    "mmap.h/mmap_overview.h",
    "hsi_combiner_guide/parameters.md",
}

POSTBUILD_TOKEN_CHECKS = {
    "NT51920": (
        "postbuild/PostbuildSetup_51920_1.3.1.bat",
        [
            r"Combiner.exe CRC_Enable output\nt51920_fw.bin",
            r"BIN\Normal_Ctrlram.bin 0x0 0x22780 10240",
            r"BIN\Normal_Ctrlram_S.bin 0x0 0x26780 10240",
            r"BIN\Vector_Ctrlram.bin 0x0 0x2D728 600",
        ],
    ),
    "NT51923": (
        "postbuild/PostbuildSetup_51923_1.4.1.bat",
        [
            r"Combiner.exe CRC_Enable output\nt51923_fw.bin",
            r"BIN\%DiffDLM% 0x0 0x28800 3072",
            r"BIN\%DiffDLM% 0x1400 0x29400 3072",
            r"BIN\%NF_Ctrlram% 0x0 0x2A000 17584",
        ],
    ),
    "NT51926": (
        "postbuild/PostbuildSetup_51926_2.0.0.bat",
        [
            r"Combiner.exe CRC_Enable output\nt51926_fw.bin",
            r"set CtrlramCMD=BIN\%Normal_Ctrlram% 0x0 %ctrlramAddr% %ctrlram_sz%",
            r"set DiffCtrlramCMD=BIN\%DiffDLM% 0x0 %diffctrlramAddr% %diffctrlram_sz%",
            r"set HeaderCMD=output\nt51926_fw.bin 0x0 %headercpyAddr% %headercpy_sz%",
        ],
    ),
    "NT51927": (
        "postbuild/PostbuildSetup_51927_1.4.1.bat",
        [
            r"Combiner.exe MERGE_MODE output\nt51927_fw.bin",
            r"BIN\%Normal_Ctrlram_M% 0x0 0x177d0 12288",
            r"BIN\%Normal_Ctrlram_SR% 0x0 0x207d0 12288",
            r"NT51927BASED_GEN_CRC_MODE CRC32 output\nt51927_fw.bin",
        ],
    ),
    "NT51930": (
        "postbuild/PostbuildSetup_51930_2.0.0.bat",
        [
            r"Combiner.exe NT51930BASED_NORMAL_MODE CRC8 output\NT51930_fw.bin output\NT51930_fw.bin",
            r"set DiffCtrlramCMD=BIN\%DiffDLM% 0x0 %diffctrlramAddr% %diffctrlram_sz%",
            r"set HeaderCMD=output\NT51930_fw.bin 0x7000 %headercpyAddr% %headercpy_sz%",
        ],
    ),
    "NT51931": (
        "postbuild/PostbuildSetup_51931_1.3.0.bat",
        [
            r"Combiner.exe NT51930BASED_NORMAL_MODE CRC8 output\nt51931_fw.bin output\nt51931_fw.bin",
            r"BIN\Normal_Ctrlram.bin 0x0 0x177D0 10240",
            r"BIN\DiffDLM.bin 0x0 0x22800 97280",
        ],
    ),
    "NT51932": (
        "postbuild/PostbuildSetup_51932_2.0.0.bat",
        [
            r"Combiner.exe NT51932BASED_NORMAL_MODE CRC8 output\nt51932_fw.bin output\nt51932_fw.bin",
            r"set NFCtrlramCMD=BIN\%NF_Ctrlram% 0x0 %nfctrlramAddr% %nfctrlram_sz%",
            r"set HeaderCMD=output\nt51932_fw.bin %headerAddr% %headercpyAddr% %headercpy_sz%",
        ],
    ),
    "NT51929": (
        "postbuild/PostbuildSetup_51932_2.0.0.bat",
        [
            r"Combiner.exe NT51932BASED_NORMAL_MODE CRC8 output\nt51932_fw.bin output\nt51932_fw.bin",
            r"set CtrlramCMD=BIN\%Normal_Ctrlram% 0x0 %ctrlramAddr% %ctrlram_sz%",
        ],
    ),
    "NT51950": (
        "postbuild/PostbuildSetup_51950_2.0.0.bat",
        [
            r"Combiner.exe NT51950BASED_NORMAL_MODE CRC8 output\nt51950_fw.bin output\nt51950_fw.bin",
            r"set DiffCtrlramCMD=BIN\%DiffDLM% 0x0 %diffctrlramAddr% %diffctrlram_sz%",
            r"set HeaderCMD=output\nt51950_fw.bin %headerAddr% %headercpyAddr% %headercpy_sz%",
        ],
    ),
    "NT51951": (
        "postbuild/PostbuildSetup_51950_2.0.0.bat",
        [
            r"Combiner.exe NT51950BASED_NORMAL_MODE CRC8 output\nt51950_fw.bin output\nt51950_fw.bin",
            r"set CtrlramCMD=BIN\%Normal_Ctrlram% 0x0 %ctrlramAddr% %ctrlram_sz%",
        ],
    ),
}

MMAP_TOKEN_CHECKS = {
    "NT51926": (
        "mmap.h/51926_2.0.0_mmap.h",
        [
            "FLASHMAP_CTRLRAM (0x22800)",
            "FLASHMAP_DIFF_DLM (0x27800)",
            "FLASHMAP_NF_TABLE (0x2C800)",
            "FLASHMAP_VNCTRLRAM (0x315D0)",
        ],
    ),
    "NT51929_NT51932": (
        "mmap.h/51929&51932_2.0.0_mmap.h",
        [
            "FLASHMAP_FW_REGISTER (0x1F200)",
            "FLASHMAP_NF_TABLE (0x1FC00)",
            "FLASHMAP_CTRLRAM (0x21B90)",
            "FLASHMAP_HEADER_COPY (0x27EF0)",
            "FLASHMAP_DIFF_DLM (0x2D100)",
        ],
    ),
    "NT51930": (
        "mmap.h/51930_2.0.0_mmap.h",
        [
            "FLASHMAP_FW_REGISTER (0x1F200)",
            "FLASHMAP_NF_TABLE (0x1FC00)",
            "FLASHMAP_CTRLRAM (0x21650)",
            "FLASHMAP_HEADER_COPY (0x28FB0)",
            "FLASHMAP_DIFF_DLM (0x2F200)",
        ],
    ),
    "NT51950": (
        "mmap.h/51950_2.0.0_mmap.h",
        [
            "FLASHMAP_FW_REGISTER (0x22200)",
            "FLASHMAP_NF_TABLE (0x22C00)",
            "FLASHMAP_CTRLRAM (0x25610)",
            "FLASHMAP_HEADER_COPY (0x2D30C)",
            "FLASHMAP_DIFF_DLM (0x33200)",
        ],
    ),
}

DOCUMENTATION_WARNINGS = [
    "NT51929/NT51932 TP Overview appears to place NF at 0x1F200, while postbuild/mmap identify FW_REGISTER=0x1F200 and NF_TABLE=0x1FC00.",
    "NT51930 TP Overview has a smaller <13 IC DiffDLM row; postbuild/mmap cascade DiffDLM is 0x2F200 size 143360.",
    "NT51927 is special: postbuild uses MERGE_MODE plus NT51927BASED_GEN_CRC_MODE CRC32; catalog includes the postbuild sequence, while mmap/TP Overview offsets still need documentation cleanup.",
    "NT51923 postbuild copies FW Config as one 2048-byte block while TP Overview splits FW Config/FW Register detail.",
    "NT51928 has no dedicated postbuild file in this archive. NT51929 follows the NT51932 reference flow and NT51951 follows the NT51950 reference flow by owner confirmation.",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().lower()


def read_compact(path: Path) -> str:
    return compact(path.read_text(encoding="utf-8", errors="replace"))


def validate_required_files(root: Path, errors: list[str]) -> None:
    for relative in sorted(REQUIRED_FILES):
        if not (root / relative).is_file():
            errors.append(f"missing reference file: {relative}")


def validate_workbook(root: Path, errors: list[str]) -> None:
    workbook_path = root / "IC_FlashMap.xlsx"
    if not workbook_path.is_file():
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append("openpyxl is required to validate IC_FlashMap.xlsx")
        return

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet_names = set(workbook.sheetnames)
    missing = REQUIRED_SHEETS - sheet_names
    if missing:
        errors.append(f"missing workbook sheets: {sorted(missing)}")

    overview = workbook["TP Overview"] if "TP Overview" in sheet_names else None
    if overview is not None:
        overview_text = " ".join(
            str(cell.value).strip()
            for row in overview.iter_rows()
            for cell in row
            if cell.value is not None
        )
        for ic in ("51920", "51923", "51926", "51927", "51928", "51929", "51930", "51931", "51932", "51950", "51951"):
            if ic not in overview_text:
                errors.append(f"TP Overview does not mention IC {ic}")


def validate_token_checks(root: Path, checks: dict[str, tuple[str, list[str]]], label: str, errors: list[str]) -> None:
    for ic, (relative, tokens) in checks.items():
        path = root / relative
        if not path.is_file():
            errors.append(f"{label} missing for {ic}: {relative}")
            continue
        text = read_compact(path)
        for token in tokens:
            if compact(token) not in text:
                errors.append(f"{label} token missing for {ic} in {relative}: {token}")


def validate_combiner(root: Path, errors: list[str]) -> None:
    path = root / "combiner_1.13.0" / "Combiner.exe"
    if path.is_file() and sha256(path) != COMBINER_SHA256:
        errors.append("Combiner.exe 1.13.0 SHA-256 does not match the approved manifest value")


def parse_args(argv: Iterable[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=DEFAULT_FLASHMAP_ROOT)
    parser.add_argument("--strict", action="store_true", help="Return non-zero when documentation warnings exist.")
    return parser.parse_args(argv)


def main(argv: Iterable[str] = sys.argv[1:]) -> int:
    args = parse_args(argv)
    root = args.root.resolve()
    errors: list[str] = []
    if not root.is_dir():
        errors.append(f"flashmap root does not exist: {root}")
    else:
        validate_required_files(root, errors)
        validate_workbook(root, errors)
        validate_token_checks(root, POSTBUILD_TOKEN_CHECKS, "postbuild", errors)
        validate_token_checks(root, MMAP_TOKEN_CHECKS, "mmap", errors)
        validate_combiner(root, errors)

    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print("IC FlashMap reference validation passed.")
    print("Verified postbuild-backed ICs: " + ", ".join(sorted(POSTBUILD_TOKEN_CHECKS)))
    print(f"Documentation warnings: {len(DOCUMENTATION_WARNINGS)}")
    for warning in DOCUMENTATION_WARNINGS:
        print(f"WARNING: {warning}")

    if args.strict and DOCUMENTATION_WARNINGS:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
